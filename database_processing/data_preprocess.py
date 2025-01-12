import os
import fireducks.pandas as pd
import glob
import logging
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

# Define expected columns
EXPECTED_COLUMNS = {
    'DataSourceName',
    'Dimension Value',
    'Owner',
    'NCM',
    'Year',
    'Month'
}

# Define constants
TARGET_SOURCE = "Base Brasil - Logcomex"

def validate_columns(df):
    """
    Validates if the DataFrame has the expected columns.
    Returns tuple of (bool, list of missing columns, list of extra columns)
    """
    current_columns = set(df.columns)
    missing_columns = EXPECTED_COLUMNS - current_columns
    extra_columns = current_columns - EXPECTED_COLUMNS
    
    return (len(missing_columns) == 0, missing_columns, extra_columns)

def filter_and_save_data(df, filepath):
    """
    1. Filters the DataFrame to keep only rows with TARGET_SOURCE
    2. Removes specified columns
    3. Reorganizes columns: Owner, NCM, Empty, Dimension Value
    4. Saves the processed data back to CSV
    Returns the filtered DataFrame and the number of rows removed.
    """
    initial_rows = len(df)
    
    # Filter for target source
    df_filtered = df[df['DataSourceName'] == TARGET_SOURCE].copy()
    rows_removed = initial_rows - len(df_filtered)
    logger.info(f"Filtered data for {TARGET_SOURCE}")
    logger.info(f"Removed {rows_removed} rows with other data sources")
    logger.info(f"Remaining rows: {len(df_filtered)}")
    
    # Remove specified columns
    columns_to_remove = ['DataSourceName', 'Year', 'Month']
    df_filtered.drop(columns=columns_to_remove, inplace=True)
    logger.info(f"Removed columns: {', '.join(columns_to_remove)}")
    
    # Reorganize columns
    # First, ensure required columns exist
    required_columns = ['Owner', 'NCM', 'Dimension Value']
    if not all(col in df_filtered.columns for col in required_columns):
        missing = [col for col in required_columns if col not in df_filtered.columns]
        raise ValueError(f"Missing required columns for reorganization: {missing}")
    
    # Keep only the required columns in the specified order
    df_filtered = df_filtered[required_columns]
    
    # Reorder columns and add empty column
    df_filtered = df_filtered[['Owner', 'NCM']]
    df_filtered.insert(2, '', '')  # Add empty column with blank name in third position
    df_filtered['Dimension Value'] = df[['Dimension Value']]  # Add Dimension Value as fourth column
    
    logger.info(f"Final column order: {', '.join(df_filtered.columns)}")
    
    # Save filtered data to CSV
    df_filtered.to_csv('Dimensions_not_registered.csv', index=False, sep=';', encoding='utf-8-sig')
    logger.info(f"Filtered data saved to 'Dimensions_not_registered.csv'")
    
    return df_filtered, rows_removed

def process_excel_file():
    """
    Checks Excel files starting with 'Dimensions' for a 'TRADEMARK' worksheet.
    If found, deletes all other worksheets, validates columns, and filters data.
    Returns a tuple of (filename, bool) indicating if the worksheet exists and was processed.
    """
    # Find all Excel files starting with 'Dimensions'
    excel_files = glob.glob('Dimensions_not_registered*.xlsx')
    
    if not excel_files:
        logger.info("No Excel files starting with 'Dimensions not registered' found in the current directory")
        return None, False
    
    # Use the most recent file if multiple exist
    if len(excel_files) > 1:
        excel_files.sort(key=os.path.getmtime, reverse=True)
        logger.info(f"Multiple files found. Using most recent: {excel_files[0]}")
    
    target_file = excel_files[0]
    
    try:
        # First check if TRADEMARK exists using pandas (faster initial check)
        xl = pd.ExcelFile(target_file)
        has_trademark = 'TRADEMARK' in xl.sheet_names
        
        if not has_trademark:
            logger.info(f"No TRADEMARK worksheet found in {target_file}")
            logger.info(f"Available worksheets: {xl.sheet_names}")
            return target_file, False
        
        # If TRADEMARK exists, use openpyxl to modify the workbook
        logger.info(f"Found TRADEMARK worksheet in {target_file}. Processing...")
        workbook = load_workbook(target_file)
        
        # Get list of all worksheets except TRADEMARK
        sheets_to_remove = [sheet for sheet in workbook.sheetnames if sheet != 'TRADEMARK']
        
        # Delete other worksheets
        for sheet_name in sheets_to_remove:
            logger.info(f"Removing worksheet: {sheet_name}")
            del workbook[sheet_name]
        
        # Save the modified workbook
        workbook.save(target_file)
        logger.info(f"Successfully processed {target_file}. Only TRADEMARK worksheet remains.")
        
        # Validate columns in the TRADEMARK worksheet
        df = pd.read_excel(target_file, sheet_name='TRADEMARK')
        is_valid, missing_cols, extra_cols = validate_columns(df)
        
        if not is_valid:
            logger.warning("Column validation failed!")
            if missing_cols:
                logger.warning(f"Missing required columns: {', '.join(missing_cols)}")
            if extra_cols:
                logger.info(f"Extra columns found: {', '.join(extra_cols)}")
            return target_file, False
        
        logger.info("All required columns are present in the TRADEMARK worksheet")
        logger.info(f"Current columns in worksheet: {', '.join(df.columns)}")
        
        # Filter and save data
        df_filtered, rows_removed = filter_and_save_data(df, target_file)
        
        return target_file, True
        
    except Exception as e:
        logger.error(f"Error processing {target_file}: {str(e)}")
        return target_file, False

if __name__ == "__main__":
    filename, success = process_excel_file()